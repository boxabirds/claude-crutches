#!/usr/bin/env python3
"""Convert an Excel file to a folder of CSVs, one per sheet.

Usage: excel_to_csv.py <path-to-excel-file>

Output: Creates <path-to-excel-file>-csv/ with one CSV per sheet.
Exits non-zero if output directory already exists or conversion fails.
"""

import csv
import os
import sys

from openpyxl import load_workbook


def sanitise_sheet_name(name: str) -> str:
    """Make a sheet name safe for use as a filename."""
    # Replace characters that are problematic in filenames
    for ch in "/\\:*?\"<>|":
        name = name.replace(ch, "_")
    return name.strip() or "unnamed"


def convert(xlsx_path: str) -> None:
    output_dir = xlsx_path + "-csv"

    if os.path.exists(output_dir):
        print(f"Output already exists: {output_dir}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    os.makedirs(output_dir)

    sheet_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sanitise_sheet_name(sheet_name)
        csv_path = os.path.join(output_dir, f"{safe_name}.csv")

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        sheet_count += 1

    wb.close()
    print(f"Converted {sheet_count} sheet(s) to {output_dir}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <excel-file>", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.isfile(path):
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)

    convert(path)
